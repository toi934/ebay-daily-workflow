"""ワンオフスクリプト（2026/07/14）

7/14朝の一括移動確認ダイアログバグ発覚を受けて、戸井さんの指示でCPaSS上で
手動処理した6件のDHL送料を、売上管理表(Google Drive上のxlsm)のBL列（通常）/
BR列（専門）に書き込む。

安全設計:
- 注文番号は列B全体をスキャンして一致する行だけを対象にする（固定行番号に依存しない）
- 該当セルが既に値ありの場合は上書きしない
- 対象外の行・セルには一切触れない
"""

import sys
import os
import tempfile

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    from openpyxl.descriptors.base import Typed as _OpxlTyped
    _orig_typed_set = _OpxlTyped.__set__
    def _patched_typed_set(self, instance, value):
        try:
            _orig_typed_set(self, instance, value)
        except (ValueError, TypeError):
            pass
    _OpxlTyped.__set__ = _patched_typed_set
except Exception:
    pass

import openpyxl
from daily_workflow_ga import (
    _get_drive_service, _find_folder_id, _list_xlsm_files,
    _download_xlsm, _upload_xlsm, find_sheet_with_orders,
    get_shipping_col, DRIVE_FOLDER_NAME,
)

# (prefix, order_no, price)
TARGETS = [
    ("通常", "03-14903-68928", 5534),
    ("通常", "11-14888-77732", 2852),
    ("通常", "11-14891-68778", 4189),
    ("通常", "15-14881-61704", 4686),
    ("通常", "19-14875-73892", 3757),
    ("専門", "24-14881-56136", 17892),
]


def process_file(service, folder_id, prefix, targets, workdir, dry_run=False):
    print(f"\n=== {prefix} ===")
    files = _list_xlsm_files(service, folder_id, prefix)
    if not files:
        print(f"  ファイルなし")
        return
    latest = files[0]
    local_path = os.path.join(workdir, latest["name"])
    _download_xlsm(service, latest["id"], local_path)
    print(f"  DL: {latest['name']} (id={latest['id']})")

    shipping_col = get_shipping_col(local_path)
    ship_col_idx = openpyxl.utils.column_index_from_string(shipping_col)

    wb = openpyxl.load_workbook(local_path, keep_vba=True)
    sheet_name = find_sheet_with_orders(wb)
    ws = wb[sheet_name]
    print(f"  対象シート: {sheet_name}")

    # 列B全体をスキャンして注文番号→行番号のマップを作る
    order_to_row = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, str):
            order_to_row[val.strip()] = row

    changed = False
    for (_, order_no, price) in targets:
        row = order_to_row.get(order_no)
        if row is None:
            print(f"  [NOT FOUND] {order_no} が列Bに見つかりません → スキップ")
            continue
        cell = ws.cell(row=row, column=ship_col_idx)
        before = cell.value
        if before is not None and before != "":
            print(f"  [SKIP] 行{row} {order_no}: {shipping_col}列は既に値あり(before={before!r}) → 上書きしない")
            continue
        print(f"  行{row} {order_no} {shipping_col}: 空欄 → {price} に記入")
        if not dry_run:
            cell.value = price
            changed = True

    if dry_run:
        print("  [DRY RUN] 保存・アップロードなし")
        wb.close()
        return

    if changed:
        wb.save(local_path)
        wb.close()
        _upload_xlsm(service, latest["id"], local_path)
        print(f"  保存・アップロード完了: {latest['name']}")
    else:
        wb.close()
        print("  変更なし（アップロードなし）")


def main():
    dry_run = "--dry-run" in sys.argv
    print(f"モード: {'DRY RUN' if dry_run else '本番'}")
    service = _get_drive_service()
    folder_id = _find_folder_id(service, DRIVE_FOLDER_NAME)
    print(f"フォルダID: {folder_id}")
    with tempfile.TemporaryDirectory() as workdir:
        for prefix in ["通常", "専門"]:
            targets = [t for t in TARGETS if t[0] == prefix]
            if not targets:
                continue
            process_file(service, folder_id, prefix, targets, workdir, dry_run=dry_run)
    print("\n完了")


if __name__ == "__main__":
    main()
