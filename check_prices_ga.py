"""読み取り専用の確認スクリプト（2026/07/09）
対象11件のBL/BR列の現在値を報告するだけ。書き込みは一切行わない。
"""
import os
import sys
import json

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

SCOPES = ["https://www.googleapis.com/auth/drive"]
DRIVE_FOLDER_NAME = "売上管理表"

sa_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
info = json.loads(sa_json)
if "private_key" in info:
    info["private_key"] = info["private_key"].replace(chr(92) + chr(110), chr(10))
creds = Credentials.from_service_account_info(info, scopes=SCOPES)
service = build("drive", "v3", credentials=creds)


def find_folder_id(name):
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    resp = service.files().list(q=q, fields="files(id,name)").execute()
    files = resp.get("files", [])
    if not files:
        raise RuntimeError("folder not found: " + name)
    return files[0]["id"]


def list_xlsm(folder_id, prefix):
    q = f"'{folder_id}' in parents and name contains '{prefix}' and name contains '.xlsm' and trashed=false"
    resp = service.files().list(q=q, fields="files(id,name,modifiedTime)", orderBy="modifiedTime desc").execute()
    return resp.get("files", [])


folder_id = find_folder_id(DRIVE_FOLDER_NAME)
print("folder_id:", folder_id)

targets = {
    "通常": {"col": "BL", "orders": ["02-14885-63234", "06-14878-91869", "06-14879-84362",
                                       "08-14873-31181", "10-14870-71397", "10-14870-90100",
                                       "15-14861-06648", "22-14851-69632"]},
    "専門": {"col": "BR", "orders": ["20-14853-54694", "09-14873-02935", "27-14843-81350"]},
}

for prefix, cfg in targets.items():
    files = list_xlsm(folder_id, prefix)
    if not files:
        print(f"[ERROR] {prefix}: no files found")
        continue
    f = files[0]
    print(f"\n=== {prefix}: {f['name']} (id={f['id']}, modified={f['modifiedTime']}) ===")
    local_path = f"/tmp/{prefix}_check.xlsm"
    request = service.files().get_media(fileId=f["id"])
    with open(local_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

    wb = openpyxl.load_workbook(local_path, keep_vba=True, data_only=True)
    ws = wb.active
    col_letter = cfg["col"]
    found = []
    missing = set(cfg["orders"])
    for row in range(2, ws.max_row + 1):
        order_val = ws.cell(row=row, column=2).value
        if order_val in cfg["orders"]:
            cell_val = ws[f"{col_letter}{row}"].value
            found.append((order_val, row, cell_val))
            missing.discard(order_val)
    for order_val, row, cell_val in found:
        status = "★空欄のまま" if cell_val is None else ("★★10078のまま(未修正)" if cell_val == 10078 else f"取得済み: {cell_val}")
        print(f"  row={row} order={order_val} {col_letter}={cell_val!r}  -> {status}")
    if missing:
        print(f"  [WARN] シートに見つからない注文: {missing}")

print("\nDONE")
