"""ワンオフ修正スクリプト（2026/07/13）

run#74で、cpass_workflow.py の DHL価格取得ロジックのバグにより、
19件（通常12件・専門7件）全てに同一の誤った価格「4687円」が
BL/BR列に書き込まれてしまった。このスクリプトはその誤データを
空欄に戻す（bug④ 2026/07/09と同じ復旧パターン）。

対象: 最新の 通常*.xlsm / 専門*.xlsm の BL列・BR列で値が 4687 のセルを
      全て None（空欄）に戻して保存・Google Driveへ再アップロードする。
"""

import sys
import os
import tempfile
import json

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    from openpyxl.descriptors.base import Typed as _OpxlTyped
    _orig_typed_set = _OpxlTyped.__set__
    def _patched_typed_set(self, instance, value):
        try:
            _orig_typed_set(self, instance, value)
        except (ValueError, TypeError):
            pass
    _OpxlTyped.__set__ = _patched_typed_set
except Exception:
    pass

import openpyxl
from daily_workflow_ga import (
    _get_drive_service, _find_folder_id, _list_xlsm_files,
    _download_xlsm, _upload_xlsm, find_sheet_with_orders,
    get_shipping_col, DRIVE_FOLDER_NAME,
)

WRONG_PRICE = 4687


def fix_file(service, folder_id, prefix, workdir, dry_run=False):
    files = _list_xlsm_files(service, folder_id, prefix)
    if not files:
        print(f"{prefix}: ファイルなし")
        return
    latest = files[0]
    local_path = os.path.join(workdir, latest["name"])
    _download_xlsm(service, latest["id"], local_path)
    print(f"DL: {latest['name']} (id={latest['id']})")

    shipping_col = get_shipping_col(local_path)
    ship_col_idx = openpyxl.utils.column_index_from_string(shipping_col)

    wb = openpyxl.load_workbook(local_path, keep_vba=True)
    sheet_name = find_sheet_with_orders(wb)
    ws = wb[sheet_name]

    fixed = []
    for row in range(2, (ws.max_row or 9999) + 1):
        order_val = ws.cell(row=row, column=2).value
        if not order_val or not isinstance(order_val, str):
            continue
        cell = ws.cell(row=row, column=ship_col_idx)
        v = cell.value
        try:
            is_wrong = v is not None and int(v) == WRONG_PRICE
        except (TypeError, ValueError):
            is_wrong = False
        if is_wrong:
            fixed.append((row, order_val.strip(), v))
            if not dry_run:
                cell.value = None

    print(f"{prefix}: {shipping_col}列 {WRONG_PRICE}円 該当 {len(fixed)}件")
    for row, order_no, v in fixed:
        print(f"  行{row} {order_no} {shipping_col}={v!r} → 空欄化")

    if dry_run:
        print("  [DRY RUN] 保存・アップロードなし")
        wb.close()
        return

    if fixed:
        wb.save(local_path)
        wb.close()
        _upload_xlsm(service, latest["id"], local_path)
        print(f"  保存・アップロード完了: {latest['name']}")
    else:
        wb.close()
        print("  修正対象なし（アップロードなし）")


def main():
    dry_run = "--dry-run" in sys.argv
    service = _get_drive_service()
    folder_id = _find_folder_id(service, DRIVE_FOLDER_NAME)
    print(f"フォルダID: {folder_id}")
    with tempfile.TemporaryDirectory() as workdir:
        for prefix in ["通常", "専門"]:
            fix_file(service, folder_id, prefix, workdir, dry_run=dry_run)
    print("完了")


if __name__ == "__main__":
    main()
