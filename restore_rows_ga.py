"""ワンオフ復旧スクリプト（2026/07/13）

fix_wrong_prices_ga.py が「BL/BR列の値が4687円のセル」を列全体から検索して
空欄化したところ、本日(7/13)の誤データ19件だけでなく、無関係な過去の正規
データ3件（たまたま元々4687円だった注文）まで巻き込んで空欄にしてしまった。

このスクリプトは、行番号を明示的に指定して、その3件だけを元の4687円に戻す
（値ベースではなく行番号ベースで復旧することで、再び余計な行を触らないようにする）。

対象:
  通常_*.xlsm  行1217 (20-12806-50263) BL列 → 4687
  通常_*.xlsm  行4758 (27-14778-47232) BL列 → 4687
  専門_*.xlsm  行502  (10-14770-62072) BR列 → 4687
"""

import sys
import os
import tempfile
import json

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

try:
    from openpyxl.descriptors.base import Typed as _OpxlTyped
    _orig_typed_set = _OpxlTyped.__set__
    def _patched_typed_set(self, instance, value):
        try:
            _orig_typed_set(self, instance, value)
        except (ValueError, TypeError):
            pass
    _OpxlTyped.__set__ = _patched_typed_set
except Exception:
    pass

import openpyxl
from daily_workflow_ga import (
    _get_drive_service, _find_folder_id, _list_xlsm_files,
    _download_xlsm, _upload_xlsm, find_sheet_with_orders,
    get_shipping_col, DRIVE_FOLDER_NAME,
)

# (prefix, row, expected_order_no, restore_value)
RESTORE_TARGETS = [
    ("通常", 1217, "20-12806-50263", 4687),
    ("通常", 4758, "27-14778-47232", 4687),
    ("専門", 502, "10-14770-62072", 4687),
]


def restore_file(service, folder_id, prefix, targets, workdir, dry_run=False):
    files = _list_xlsm_files(service, folder_id, prefix)
    if not files:
        print(f"{prefix}: ファイルなし")
        return
    latest = files[0]
    local_path = os.path.join(workdir, latest["name"])
    _download_xlsm(service, latest["id"], local_path)
    print(f"DL: {latest['name']} (id={latest['id']})")

    shipping_col = get_shipping_col(local_path)
    ship_col_idx = openpyxl.utils.column_index_from_string(shipping_col)

    wb = openpyxl.load_workbook(local_path, keep_vba=True)
    sheet_name = find_sheet_with_orders(wb)
    ws = wb[sheet_name]

    done = []
    for (_, row, expected_order, value) in targets:
        order_val = ws.cell(row=row, column=2).value
        order_val = order_val.strip() if isinstance(order_val, str) else order_val
        if order_val != expected_order:
            print(f"  [SKIP] 行{row}: 注文番号不一致 (期待={expected_order} 実際={order_val!r}) → 触らない")
            continue
        cell = ws.cell(row=row, column=ship_col_idx)
        before = cell.value
        if before is not None:
            print(f"  [SKIP] 行{row} {expected_order}: {shipping_col}列は既に値あり(before={before!r}) → 上書きしない")
            continue
        done.append((row, expected_order, value))
        if not dry_run:
            cell.value = value
        print(f"  行{row} {expected_order} {shipping_col}: 空欄 → {value} に復旧")

    if dry_run:
        print("  [DRY RUN] 保存・アップロードなし")
        wb.close()
        return

    if done:
        wb.save(local_path)
        wb.close()
        _upload_xlsm(service, latest["id"], local_path)
        print(f"  保存・アップロード完了: {latest['name']}")
    else:
        wb.close()
        print("  復旧対象なし（アップロードなし）")


def main():
    dry_run = "--dry-run" in sys.argv
    service = _get_drive_service()
    folder_id = _find_folder_id(service, DRIVE_FOLDER_NAME)
    print(f"フォルダID: {folder_id}")
    with tempfile.TemporaryDirectory() as workdir:
        for prefix in ["通常", "専門"]:
            targets = [t for t in RESTORE_TARGETS if t[0] == prefix]
            if not targets:
                continue
            restore_file(service, folder_id, prefix, targets, workdir, dry_run=dry_run)
    print("完了")


if __name__ == "__main__":
    main()
