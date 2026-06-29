"""毎朝11:00 売上管理表ワークフロー (GitHub Actions版)

ローカル版との違い:
- Google Drive APIでxlsmをDL/UL（ローカルマイドライブ不要）
- win32comの代わりにopenpyxlで書き込み（keep_vba=Trueで保護）
- CPaSS/Playwrightはheadlessモードで実行
- 完了後にGmailでメール送信

必要なGitHub Secrets:
  CPASS_EMAIL / CPASS_PASSWORD
  GOOGLE_SERVICE_ACCOUNT_JSON
  GMAIL_APP_PASSWORD
"""

import sys
import os
import re
import time
import json
import io
import smtplib
import tempfile
from email.mime.text import MIMEText
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# openpyxl互換バグ回避
try:
    from openpyxl.descriptors.base import Typed as _OpxlTyped
    _orig_typed_set = _OpxlTyped.__set__
    def _patched_typed_set(self, instance, value):
        try:
            _orig_typed_set(self, instance, value)
        except (ValueError, TypeError):
            pass
    _OpxlTyped.__set__ = _patched_typed_set
except Exception:
    pass

import openpyxl
import cpass_workflow

# ─── 定数 ───
E_COL_VALUE = "③マーキング番号、リサーチ者記入"
F_COL_VALUE = "仕入未"
EXCHANGE_RATE_KEYWORDS = ["為替", "為"]
DRIVE_FOLDER_NAME = "売上管理表"
GMAIL_FROM = "gen7m9@gmail.com"


# ─── Google Drive API ───
def _get_drive_service():
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    sa_json = os.environ["GOOGLE_SERVICE_ACCOUNT_JSON"]
    info = json.loads(sa_json)
    # GitHub Secretsの改行エスケープ問題を修正（\\n → 実改行）
    if "private_key" in info:
        info["private_key"] = info["private_key"].replace("\\\\n", "\\n")
    creds = Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)


def _find_folder_id(service, folder_name):
    """マイドライブ内のフォルダIDを取得"""
    q = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    resp = service.files().list(q=q, fields="files(id,name)").execute()
    files = resp.get("files", [])
    if not files:
        raise RuntimeError(f"Google Driveにフォルダ '{folder_name}' が見つかりません")
    return files[0]["id"]


def _list_xlsm_files(service, folder_id, prefix):
    """フォルダ内のxlsmファイル一覧（prefix: '通常' or '専門'）を取得"""
    q = f"'{folder_id}' in parents and name contains '{prefix}' and name contains '.xlsm' and trashed=false"
    resp = service.files().list(q=q, fields="files(id,name,modifiedTime)", orderBy="modifiedTime desc").execute()
    return resp.get("files", [])


def _download_xlsm(service, file_id, dest_path):
    """Google DriveからxlsmをDL"""
    from googleapiclient.http import MediaIoBaseDownload
    request = service.files().get_media(fileId=file_id)
    with open(dest_path, "wb") as f:
        downloader = MediaIoBaseDownload(f, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()


def _upload_xlsm(service, file_id, src_path):
    """Google Driveの既存ファイルを上書きUL"""
    from googleapiclient.http import MediaFileUpload
    media = MediaFileUpload(
        src_path,
        mimetype="application/vnd.ms-excel.sheet.macroEnabled.12",
        resumable=False
    )
    service.files().update(fileId=file_id, media_body=media).execute()
    print(f"  Google Driveにアップロード完了: {os.path.basename(src_path)}")


def download_xlsm_files(workdir):
    """通常・専門の最新xlsmをDL。{prefix: (local_path, drive_file_id)} を返す"""
    print("=" * 60)
    print("Step 1: Google DriveからDL")
    print("=" * 60)
    service = _get_drive_service()
    folder_id = _find_folder_id(service, DRIVE_FOLDER_NAME)
    print(f"  フォルダID: {folder_id}")

    result = {}
    for prefix in ["通常", "専門"]:
        files = _list_xlsm_files(service, folder_id, prefix)
        if not files:
            print(f"  {prefix}: ファイルなし")
            continue
        latest = files[0]  # modifiedTime descでソート済み
        local_path = os.path.join(workdir, latest["name"])
        _download_xlsm(service, latest["id"], local_path)
        print(f"  DL: {latest['name']} (id={latest['id']})")
        result[prefix] = (local_path, latest["id"])

    return service, folder_id, result


# ─── Excel処理 ───
def get_shipping_col(xlsm_path):
    if "通常" in os.path.basename(xlsm_path):
        return "BL"
    return "BR"


def find_sheet_with_orders(wb):
    from openpyxl.chartsheet.chartsheet import Chartsheet
    order_pattern = re.compile(r"^\d{2}-\d{5}-\d{5}$")
    best_sheet, best_count = None, 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        if isinstance(ws, Chartsheet):
            continue
        count = 0
        for row in range(2, min(500, ws.max_row) + 1):
            val = ws.cell(row=row, column=2).value
            if val and isinstance(val, str) and order_pattern.match(val.strip()):
                count += 1
        if count > best_count:
            best_count, best_sheet = count, sname
    print(f"  対象シート: {best_sheet} (注文番号{best_count}件)")
    return best_sheet


def find_exchange_cols(ws):
    found = []
    for row in range(1, 4):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val and isinstance(val, str) and any(kw in val for kw in EXCHANGE_RATE_KEYWORDS):
                found.append({"row": row, "col": col,
                              "letter": openpyxl.utils.get_column_letter(col), "value": val})
    return found


def get_target_orders(xlsm_path):
    """BL/BR空白の新規注文番号セットを返す"""
    order_pattern = re.compile(r"^\d{2}-\d{5}-\d{5}$")
    shipping_col = get_shipping_col(xlsm_path)
    ship_col_idx = openpyxl.utils.column_index_from_string(shipping_col)

    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    sheet_name = find_sheet_with_orders(wb)
    if not sheet_name:
        wb.close()
        return set()
    ws = wb[sheet_name]

    last_filled_row = 1
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if not val or not isinstance(val, str) or not order_pattern.match(val.strip()):
            continue
        br_val = ws.cell(row=row, column=ship_col_idx).value
        if br_val is not None and br_val != "":
            last_filled_row = row

    print(f"  送料記入済み最終行: {last_filled_row}")

    targets = set()
    for row in range(last_filled_row + 1, ws.max_row + 1):
        val = ws.cell(row=row, column=2).value
        if not val or not isinstance(val, str) or not order_pattern.match(val.strip()):
            continue
        e_val = ws.cell(row=row, column=5).value
        if "キャンセル" in str(e_val or ""):
            continue
        br_val = ws.cell(row=row, column=ship_col_idx).value
        if br_val is None or br_val == "":
            targets.add(val.strip())

    wb.close()
    return targets


def process_xlsm(xlsm_path, cpass_results, dry_run=False):
    """xlsmを開いてCPaSSデータで空白セルを埋め、openpyxlで保存"""
    print()
    print("=" * 60)
    print(f"Step 3: {os.path.basename(xlsm_path)}")
    print("=" * 60)

    shipping_col = get_shipping_col(xlsm_path)
    ship_col_idx = openpyxl.utils.column_index_from_string(shipping_col)
    order_pattern = re.compile(r"^\d{2}-\d{5}-\d{5}$")

    # 読み取り専用で開いてwritesを収集
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True, data_only=True)
    sheet_name = find_sheet_with_orders(wb)
    if not sheet_name:
        print("  注文データのシートが見つかりません")
        wb.close()
        return False
    ws = wb[sheet_name]

    ex_cols = find_exchange_cols(ws)
    writes = []
    fill_details = {"A": 0, "E": 0, "F": 0, shipping_col: 0, "exchange": 0}

    for row in range(2, (ws.max_row or 9999) + 1):
        order_val = ws.cell(row=row, column=2).value
        if not order_val or not isinstance(order_val, str):
            continue
        order_no = order_val.strip()
        if not order_pattern.match(order_no):
            continue

        cpass_info = cpass_results.get(order_no)
        e_val = ws.cell(row=row, column=5).value
        is_cancel = "キャンセル" in str(e_val or "")

        if cpass_info and cpass_info.get("package_no"):
            if not ws.cell(row=row, column=1).value:
                writes.append((row, 1, cpass_info["package_no"]))
                fill_details["A"] += 1

        if not e_val:
            writes.append((row, 5, E_COL_VALUE))
            fill_details["E"] += 1

        f_val = ws.cell(row=row, column=6).value
        if not f_val:
            writes.append((row, 6, F_COL_VALUE))
            fill_details["F"] += 1

        if cpass_info and cpass_info.get("dhl_price_jpy"):
            br_val = ws.cell(row=row, column=ship_col_idx).value
            if not br_val:
                writes.append((row, ship_col_idx, int(cpass_info["dhl_price_jpy"])))
                fill_details[shipping_col] += 1

        if not is_cancel:
            for ex_col in ex_cols:
                cell_val = ws.cell(row=row, column=ex_col["col"]).value
                if cell_val is None or (isinstance(cell_val, str) and cell_val.strip() == ""):
                    if row > 2:
                        prev_val = ws.cell(row=row - 1, column=ex_col["col"]).value
                        if prev_val is not None and prev_val != "":
                            writes.append((row, ex_col["col"], prev_val))
                            fill_details["exchange"] += 1

    wb.close()

    print(f"  書き込み予定: {json.dumps(fill_details, ensure_ascii=False)}")

    if dry_run:
        print("  [DRY RUN] 書き込みなし")
        return True

    if not writes:
        print("  書き込みなし（全セル埋まり済み）")
        return True

    # openpyxlで書き込み（keep_vba=TrueでVBA保持）
    print(f"  openpyxl で書き込み中... ({len(writes)} セル)")
    wb2 = openpyxl.load_workbook(xlsm_path, keep_vba=True)
    ws2 = wb2[sheet_name]
    for (row, col, value) in writes:
        ws2.cell(row=row, column=col).value = value
    wb2.save(xlsm_path)
    wb2.close()
    print(f"  保存OK: {os.path.basename(xlsm_path)}")
    return True


# ─── メール送信 ───
def send_result_email(results_text, run_count, error_text=""):
    app_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not app_password:
        print("  GMAIL_APP_PASSWORD未設定、メールスキップ")
        return
    subject = f"タスク2（売上管理表）実行結果 ({run_count}回目)"
    body = f"実行日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}\n\n"
    body += results_text
    if error_text:
        body += f"\n\n【エラー】\n{error_text}"
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"] = GMAIL_FROM
    msg["To"] = GMAIL_FROM
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(GMAIL_FROM, app_password)
            smtp.send_message(msg)
        print(f"  メール送信: {subject}")
    except Exception as e:
        print(f"  メール送信失敗: {e}")


def _get_run_count():
    count_file = "run_count_task2.txt"
    try:
        with open(count_file) as f:
            n = int(f.read().strip()) + 1
    except Exception:
        n = 1
    with open(count_file, "w") as f:
        f.write(str(n))
    return n


# ─── メイン ───
def main():
    dry_run = "--dry-run" in sys.argv
    start_time = datetime.now()

    print("=" * 60)
    print(f"売上管理表ワークフロー (GitHub Actions版)  {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    run_count = _get_run_count()
    errors = []
    summary_lines = []

    with tempfile.TemporaryDirectory() as workdir:
        # Step 1: Google Drive からDL
        try:
            service, folder_id, dl_result = download_xlsm_files(workdir)
        except Exception as e:
            msg = f"Google Drive DL失敗: {e}"
            print(msg)
            errors.append(msg)
            send_result_email("Google Drive DL失敗", run_count, "\n".join(errors))
            return

        if not dl_result:
            print("対象ファイルなし、終了")
            send_result_email("対象ファイルなし", run_count)
            return

        # Step 2: 対象注文番号収集
        target_order_nos = set()
        xlsm_paths = {}
        for prefix, (local_path, file_id) in dl_result.items():
            try:
                targets = get_target_orders(local_path)
                target_order_nos |= targets
                xlsm_paths[prefix] = (local_path, file_id)
                print(f"  {prefix}: 対象{len(targets)}件")
                summary_lines.append(f"{prefix}: 対象注文{len(targets)}件")
            except Exception as e:
                msg = f"{prefix}の注文抽出失敗: {e}"
                print(msg)
                errors.append(msg)

        print(f"\n送料列が空白の対象注文: {len(target_order_nos)} 件")

        # Step 3: CPaSS処理
        cpass_results = {}
        if target_order_nos and not dry_run:
            print()
            print("=" * 60)
            print("Step 2: CPaSS ワークフロー実行")
            print("=" * 60)
            try:
                cpass_results = cpass_workflow.process_all_orders_for_dhl(
                    target_order_nos=target_order_nos,
                    headless=True,
                    move_waiting=True,
                )
                ok_count = sum(1 for v in cpass_results.values() if v.get("dhl_price_jpy"))
                summary_lines.append(f"CPaSS: {ok_count}/{len(target_order_nos)}件 DHL金額取得")
            except Exception as e:
                msg = f"CPaSS処理エラー: {e}"
                print(msg)
                errors.append(msg)

        # Step 4: Excelに書き込み → Google Driveにアップロード
        for prefix, (local_path, file_id) in xlsm_paths.items():
            try:
                ok = process_xlsm(local_path, cpass_results, dry_run=dry_run)
                if ok and not dry_run:
                    _upload_xlsm(service, file_id, local_path)
                    summary_lines.append(f"{prefix}: 書き込み・UL完了")
            except Exception as e:
                msg = f"{prefix}の処理失敗: {e}"
                print(msg)
                errors.append(msg)

    end_time = datetime.now()
    elapsed = int((end_time - start_time).total_seconds())

    print()
    print("=" * 60)
    print(f"完了  {end_time.strftime('%Y-%m-%d %H:%M:%S')}  ({elapsed}秒)")
    print("=" * 60)

    # メール送信
    results_text = "\n".join(summary_lines) if summary_lines else "処理完了（書き込みなし）"
    results_text += f"\n\n経過時間: {elapsed}秒"
    send_result_email(results_text, run_count, "\n".join(errors))


if __name__ == "__main__":
    main()
